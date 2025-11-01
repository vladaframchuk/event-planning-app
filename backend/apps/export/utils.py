from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any, Iterable

from django.db.models import Count, Prefetch, QuerySet
from django.utils import timezone

from apps.events.models import Event
from apps.polls.models import Poll, PollOption
from apps.tasks.models import Task, TaskList

_CSV_DELIMITER: str = ";"
_CSV_BOM: str = "\ufeff"
_DATETIME_FORMAT: str = "%d.%m.%Y %H:%M"
_DATE_FORMAT: str = "%d.%m.%Y"


@dataclass(frozen=True)
class _TaskExportRow:
    """Строка экспорта задачи с уже отформатированными полями."""

    title: str
    status: str
    assignee: str
    start_at: str
    due_at: str


@dataclass(frozen=True)
class _PollOptionExportRow:
    """Строка экспорта опроса с вариантом и числом голосов."""

    question: str
    option: str
    votes: int


def _event_base_queryset() -> QuerySet[Event]:
    """Возвращает queryset события с предвыборкой задач и опросов."""

    task_prefetch = Prefetch(
        "task_lists",
        queryset=TaskList.objects.order_by("order", "id").prefetch_related(
            Prefetch(
                "tasks",
                queryset=Task.objects.select_related("assignee__user")
                .order_by("order", "id")
                .only(
                    "id",
                    "title",
                    "status",
                    "assignee_id",
                    "assignee__user__name",
                    "assignee__user__email",
                    "start_at",
                    "due_at",
                ),
            )
        ),
    )
    poll_prefetch = Prefetch(
        "polls",
        queryset=Poll.objects.order_by("created_at", "id").prefetch_related(
            Prefetch(
                "options",
                queryset=PollOption.objects.annotate(votes_count=Count("votes")).order_by("id"),
            )
        ),
    )

    return Event.objects.only("id", "title").prefetch_related(task_prefetch, poll_prefetch)


def _format_datetime(value: Any) -> str:
    """Приводит datetime к строке в формате ДД.ММ.ГГГГ ЧЧ:ММ или возвращает тире."""

    if not value:
        return "—"
    if timezone.is_aware(value):
        localized = timezone.localtime(value)
    else:
        localized = value
    return localized.strftime(_DATETIME_FORMAT)


def _format_date(value: Any) -> str:
    """Форматирует дату опции опроса или возвращает тире."""

    if not value:
        return "—"
    return value.strftime(_DATE_FORMAT)


def _format_assignee(task: Task) -> str:
    """Возвращает имя исполнителя задачи или тире, если исполнитель отсутствует."""

    assignee = task.assignee
    if assignee is None or assignee.user is None:
        return "—"

    name = assignee.user.name
    if name:
        return name
    email = assignee.user.email
    return email or "—"


def _collect_tasks(event: Event) -> list[_TaskExportRow]:
    """Собирает и форматирует задачи события в список строк экспорта."""

    rows: list[_TaskExportRow] = []
    for task_list in event.task_lists.all():
        for task in task_list.tasks.all():
            rows.append(
                _TaskExportRow(
                    title=task.title,
                    status=task.status,
                    assignee=_format_assignee(task),
                    start_at=_format_datetime(task.start_at),
                    due_at=_format_datetime(task.due_at),
                )
            )
    return rows


def _format_poll_option(option: PollOption) -> str:
    """Форматирует вариант опроса в человеко-читаемую строку."""

    if option.label:
        return option.label
    if option.date_value:
        return _format_date(option.date_value)
    return "—"


def _collect_polls(event: Event) -> list[_PollOptionExportRow]:
    """Возвращает список строк опросов с вариантами и количеством голосов."""

    rows: list[_PollOptionExportRow] = []
    for poll in event.polls.all():
        options: Iterable[PollOption] = poll.options.all()
        has_options = False
        for index, option in enumerate(options):
            has_options = True
            votes_count = int(getattr(option, "votes_count", 0) or 0)
            rows.append(
                _PollOptionExportRow(
                    question=poll.question if index == 0 else "",
                    option=_format_poll_option(option),
                    votes=votes_count,
                )
            )
        if not has_options:
            rows.append(
                _PollOptionExportRow(
                    question=poll.question,
                    option="—",
                    votes=0,
                )
            )
    return rows


def generate_event_csv(event_id: int) -> bytes:
    """Генерирует CSV-файл по задачам и опросам события."""

    event = _event_base_queryset().get(id=event_id)
    task_rows = _collect_tasks(event)
    poll_rows = _collect_polls(event)

    buffer = StringIO(newline="")
    writer = csv.writer(buffer, delimiter=_CSV_DELIMITER)

    writer.writerow([f"Событие", event.title])
    writer.writerow([])
    writer.writerow(["Задачи"])
    writer.writerow(["Название задачи", "Статус", "Исполнитель", "Дата начала", "Дедлайн"])
    for row in task_rows:
        writer.writerow([row.title, row.status, row.assignee, row.start_at, row.due_at])

    writer.writerow([])
    writer.writerow(["Опросы"])
    writer.writerow(["Название опроса", "Вариант", "Кол-во голосов", "", ""])
    for row in poll_rows:
        writer.writerow([row.question, row.option, row.votes, "", ""])

    csv_content = buffer.getvalue()
    return (_CSV_BOM + csv_content).encode("utf-8")


def generate_event_xls(event_id: int) -> bytes:
    """Генерирует XLS-файл по задачам и опросам события."""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - выполняется только если нет openpyxl
        raise RuntimeError("openpyxl требуется для экспорта в XLS.") from exc

    event = _event_base_queryset().get(id=event_id)
    task_rows = _collect_tasks(event)
    poll_rows = _collect_polls(event)

    workbook = Workbook()
    tasks_sheet = workbook.active
    tasks_sheet.title = "Задачи"
    polls_sheet = workbook.create_sheet(title="Опросы")

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="FF1F2937")
    header_alignment = Alignment(horizontal="center", vertical="center")

    task_headers = ["Название задачи", "Статус", "Исполнитель", "Дата начала", "Дедлайн"]
    tasks_sheet.append(task_headers)
    for cell in tasks_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row in task_rows:
        tasks_sheet.append([row.title, row.status, row.assignee, row.start_at, row.due_at])

    tasks_sheet.freeze_panes = "A2"

    poll_headers = ["Название опроса", "Вариант", "Кол-во голосов"]
    polls_sheet.append(poll_headers)
    for cell in polls_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row in poll_rows:
        polls_sheet.append([row.question, row.option, row.votes])

    polls_sheet.freeze_panes = "A2"

    def _autofit(sheet: Any) -> None:
        for column_index, column_cells in enumerate(sheet.columns, start=1):
            max_length = 0
            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue
                text = str(value)
                if len(text) > max_length:
                    max_length = len(text)
            adjusted_width = max(max_length + 2, 12)
            column_letter = get_column_letter(column_index)
            sheet.column_dimensions[column_letter].width = adjusted_width

    _autofit(tasks_sheet)
    _autofit(polls_sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

